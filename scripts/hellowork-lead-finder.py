#!/usr/bin/env python3
"""
YUMEHO 見込み施設抽出ツール
ハローワークインターネットサービスから求人情報を取得し、
YUMEHO の導入ターゲットとなる施設をスコアリング・リスト化する。

使い方:
  python3 scripts/hellowork-lead-finder.py search          # 検索 → CSV出力
  python3 scripts/hellowork-lead-finder.py search --area 東京都  # エリア指定
  python3 scripts/hellowork-lead-finder.py analyze data.csv  # 既存CSVを再スコアリング
"""

import csv
import json
import os
import re
import sys
import time
from collections import defaultdict
from datetime import datetime, timedelta
from urllib.parse import urlencode, urljoin

import requests
from bs4 import BeautifulSoup

# ── 定数 ──────────────────────────────────────────

BASE_URL = "https://www.hellowork.mhlw.go.jp"
SEARCH_URL = f"{BASE_URL}/kensaku/GECA110010.do"

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "leads")

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "ja,en-US;q=0.9,en;q=0.8",
}

REQUEST_INTERVAL = 3  # 秒（サーバー負荷軽減）

# ── ターゲット条件 ──────────────────────────────────

TARGET_KEYWORDS = [
    "理学療法士",
    "作業療法士",
    "リハビリ",
    "介護職員",
    "機能訓練指導員",
    "歩行訓練",
]

TARGET_FACILITY_TYPES = [
    "回復期リハビリテーション",
    "回復期リハビリ",
    "介護老人保健施設",
    "老健",
    "通所リハビリ",
    "デイケア",
    "デイサービス",
    "リハビリテーション病院",
    "リハビリ病院",
    "地域包括ケア",
    "訪問リハビリ",
    "介護医療院",
    "特別養護老人ホーム",
    "特養",
    "有料老人ホーム",
]

HIGH_VALUE_SIGNALS = [
    "人手不足",
    "急募",
    "増員",
    "欠員補充",
    "即日",
    "随時",
    "未経験可",
    "資格不問",
    "経験不問",
    "ブランク可",
    "複数名",
    "2名",
    "3名",
    "歩行介助",
    "歩行訓練",
    "転倒",
    "転倒防止",
    "身体介助",
    "移乗",
]

AREA_CODES = {
    "北海道": "01", "青森県": "02", "岩手県": "03", "宮城県": "04",
    "秋田県": "05", "山形県": "06", "福島県": "07", "茨城県": "08",
    "栃木県": "09", "群馬県": "10", "埼玉県": "11", "千葉県": "12",
    "東京都": "13", "神奈川県": "14", "新潟県": "15", "富山県": "16",
    "石川県": "17", "福井県": "18", "山梨県": "19", "長野県": "20",
    "岐阜県": "21", "静岡県": "22", "愛知県": "23", "三重県": "24",
    "滋賀県": "25", "京都府": "26", "大阪府": "27", "兵庫県": "28",
    "奈良県": "29", "和歌山県": "30", "鳥取県": "31", "島根県": "32",
    "岡山県": "33", "広島県": "34", "山口県": "35", "徳島県": "36",
    "香川県": "37", "愛媛県": "38", "高知県": "39", "福岡県": "40",
    "佐賀県": "41", "長崎県": "42", "熊本県": "43", "大分県": "44",
    "宮崎県": "45", "鹿児島県": "46", "沖縄県": "47",
}

DEFAULT_AREAS = ["東京都", "神奈川県", "埼玉県", "千葉県", "大阪府", "愛知県"]


# ── スコアリング ──────────────────────────────────

def score_listing(listing):
    """求人情報をスコアリングして優先度を算出する。"""
    score = 0
    reasons = []
    text = " ".join([
        listing.get("job_title", ""),
        listing.get("company_name", ""),
        listing.get("description", ""),
        listing.get("requirements", ""),
    ]).lower()

    # 施設種別スコア
    for ft in TARGET_FACILITY_TYPES:
        if ft.lower() in text:
            score += 20
            reasons.append(f"施設種別: {ft}")
            break

    # 職種スコア
    job = listing.get("job_title", "")
    if re.search(r"理学療法士|PT", job):
        score += 15
        reasons.append("PT求人")
    elif re.search(r"作業療法士|OT", job):
        score += 15
        reasons.append("OT求人")
    elif re.search(r"介護職|介護員|ケアワーカー", job):
        score += 10
        reasons.append("介護職求人")
    elif re.search(r"機能訓練|リハビリ", job):
        score += 12
        reasons.append("リハビリ関連求人")

    # 高優先シグナル
    for signal in HIGH_VALUE_SIGNALS:
        if signal in text:
            score += 5
            reasons.append(f"シグナル: {signal}")

    # 歩行リハビリ直接言及
    if re.search(r"歩行.*リハビリ|歩行.*訓練|歩行.*介助", text):
        score += 25
        reasons.append("歩行リハビリ直接言及")

    # 転倒関連
    if "転倒" in text:
        score += 15
        reasons.append("転倒関連言及")

    # 人数（複数名募集 = 深刻な人手不足）
    match = re.search(r"(\d+)名.*募集|募集.*(\d+)名", text)
    if match:
        num = int(match.group(1) or match.group(2))
        if num >= 2:
            score += 10 * min(num, 5)
            reasons.append(f"複数名募集: {num}名")

    # 急募
    if "急募" in text:
        score += 15
        reasons.append("急募")

    listing["score"] = score
    listing["score_reasons"] = "; ".join(reasons[:8])
    listing["priority"] = (
        "A (最優先)" if score >= 60 else
        "B (有望)" if score >= 35 else
        "C (標準)" if score >= 15 else
        "D (低)"
    )
    return listing


# ── ハローワーク検索 ──────────────────────────────

class HelloWorkSearcher:
    """ハローワークインターネットサービスの公開求人を検索する。"""

    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update(HEADERS)
        self.results = []

    def search(self, keyword, area_name=None, max_pages=5):
        """キーワードとエリアで求人を検索する。"""
        print(f"\n  検索中: 「{keyword}」", end="")
        if area_name:
            print(f" / {area_name}", end="")
        print()

        params = {
            "screenId": "GECA110010",
            "action": "searchAction",
            "searchTarget": "1",
            "freeWordType": "1",
            "freeWord": keyword,
            "searchButton": "検索",
        }

        if area_name and area_name in AREA_CODES:
            params["tDFK1CmbBox"] = AREA_CODES[area_name]

        page_count = 0
        listings_found = 0

        try:
            resp = self.session.get(SEARCH_URL, params=params, timeout=30)
            resp.raise_for_status()

            for page in range(max_pages):
                page_count += 1
                new_listings = self._parse_search_results(resp.text)

                if not new_listings:
                    break

                listings_found += len(new_listings)
                self.results.extend(new_listings)

                print(f"    ページ {page_count}: {len(new_listings)} 件取得")

                next_url = self._find_next_page(resp.text)
                if not next_url:
                    break

                time.sleep(REQUEST_INTERVAL)
                resp = self.session.get(urljoin(BASE_URL, next_url), timeout=30)
                resp.raise_for_status()

        except requests.RequestException as e:
            print(f"    [エラー] {e}")

        print(f"    → 合計 {listings_found} 件")
        return listings_found

    def _parse_search_results(self, html):
        """検索結果ページから求人情報を抽出する。"""
        soup = BeautifulSoup(html, "html.parser")
        listings = []

        # ハローワークの求人カード要素を探索
        # 2024年以降のUIでは .kyujin-row or .job-item 等のクラスが使われる
        job_cards = soup.select(".kyujin-row, .job-item, .result-item, tr.job")

        if not job_cards:
            # フォールバック: テーブル行から抽出を試みる
            job_cards = soup.select("table.job-table tr, .search-result-item")

        if not job_cards:
            # 最終フォールバック: 求人番号パターンで探す
            job_cards = self._extract_by_pattern(soup)

        for card in job_cards:
            listing = self._parse_card(card)
            if listing and listing.get("company_name"):
                listings.append(listing)

        return listings

    def _parse_card(self, card):
        """1件の求人カードから情報を抽出する。"""
        text = card.get_text(separator=" ", strip=True)

        listing = {
            "job_number": "",
            "job_title": "",
            "company_name": "",
            "location": "",
            "salary": "",
            "description": text[:500],
            "requirements": "",
            "employment_type": "",
            "source": "ハローワーク",
            "retrieved_at": datetime.now().strftime("%Y-%m-%d"),
        }

        # 求人番号
        num_match = re.search(r"(\d{5}-\d{8,}|\d{2}\d{3}-\d{8})", text)
        if num_match:
            listing["job_number"] = num_match.group(1)

        # リンクからタイトルや企業名を取得
        links = card.select("a")
        for link in links:
            link_text = link.get_text(strip=True)
            if link_text and len(link_text) > 2:
                if not listing["job_title"]:
                    listing["job_title"] = link_text
                elif not listing["company_name"]:
                    listing["company_name"] = link_text

        # テーブルセルから情報抽出（th/td パターン）
        for th in card.select("th, dt"):
            label = th.get_text(strip=True)
            td = th.find_next_sibling("td") or th.find_next_sibling("dd")
            if not td:
                continue
            value = td.get_text(strip=True)

            if "事業所名" in label or "会社名" in label or "企業名" in label:
                listing["company_name"] = value
            elif "職種" in label:
                listing["job_title"] = value
            elif "就業場所" in label or "勤務地" in label or "所在地" in label:
                listing["location"] = value
            elif "賃金" in label or "給与" in label:
                listing["salary"] = value
            elif "仕事の内容" in label or "業務内容" in label:
                listing["description"] = value
            elif "雇用形態" in label:
                listing["employment_type"] = value

        # テキストパターンからの抽出（カード形式の場合）
        if not listing["location"]:
            loc_match = re.search(
                r"(北海道|青森|岩手|宮城|秋田|山形|福島|茨城|栃木|群馬|"
                r"埼玉|千葉|東京|神奈川|新潟|富山|石川|福井|山梨|長野|"
                r"岐阜|静岡|愛知|三重|滋賀|京都|大阪|兵庫|奈良|和歌山|"
                r"鳥取|島根|岡山|広島|山口|徳島|香川|愛媛|高知|福岡|"
                r"佐賀|長崎|熊本|大分|宮崎|鹿児島|沖縄)[都道府県]?.{0,20}",
                text
            )
            if loc_match:
                listing["location"] = loc_match.group(0)[:30]

        return listing

    def _extract_by_pattern(self, soup):
        """パターンマッチで求人情報のブロックを抽出する。"""
        blocks = []
        body_text = soup.get_text()

        # 求人番号パターンで分割
        parts = re.split(r"(?=\d{5}-\d{8,})", body_text)
        for part in parts[1:]:  # 最初の空ブロックをスキップ
            tag = BeautifulSoup(f"<div>{part[:1000]}</div>", "html.parser")
            blocks.append(tag.find("div"))

        return blocks[:50]  # 最大50件

    def _find_next_page(self, html):
        """次ページへのリンクを探す。"""
        soup = BeautifulSoup(html, "html.parser")
        next_link = soup.select_one("a.next, a[title*='次'], a:contains('次へ'), .pagination a.next")
        if next_link and next_link.get("href"):
            return next_link["href"]
        # テキストで探す
        for a in soup.select("a"):
            if "次へ" in a.get_text() or "次の" in a.get_text():
                if a.get("href"):
                    return a["href"]
        return None


# ── 施設の名寄せ・集約 ──────────────────────────────

def deduplicate_and_aggregate(listings):
    """同一施設の複数求人を名寄せし、施設単位で集約する。"""
    facilities = defaultdict(lambda: {
        "company_name": "",
        "location": "",
        "job_count": 0,
        "job_titles": [],
        "best_score": 0,
        "all_reasons": set(),
        "priority": "D (低)",
        "descriptions": [],
        "salary_range": "",
        "job_numbers": [],
        "retrieved_at": "",
    })

    for listing in listings:
        name = normalize_company_name(listing.get("company_name", ""))
        if not name:
            continue

        f = facilities[name]
        f["company_name"] = listing.get("company_name", name)
        f["location"] = f["location"] or listing.get("location", "")
        f["job_count"] += 1
        if listing.get("job_title"):
            f["job_titles"].append(listing["job_title"])
        if listing.get("job_number"):
            f["job_numbers"].append(listing["job_number"])
        if listing.get("description"):
            f["descriptions"].append(listing["description"][:200])
        f["retrieved_at"] = listing.get("retrieved_at", "")

        scored = score_listing(listing)
        if scored["score"] > f["best_score"]:
            f["best_score"] = scored["score"]
            f["priority"] = scored["priority"]
        if scored.get("score_reasons"):
            f["all_reasons"].update(scored["score_reasons"].split("; "))

    # 複数求人を出している施設にボーナス
    result = []
    for name, f in facilities.items():
        if f["job_count"] >= 3:
            f["best_score"] += 20
            f["all_reasons"].add(f"複数職種で求人: {f['job_count']}件")
        elif f["job_count"] >= 2:
            f["best_score"] += 10
            f["all_reasons"].add(f"複数求人: {f['job_count']}件")

        # 優先度再計算
        s = f["best_score"]
        f["priority"] = (
            "A (最優先)" if s >= 60 else
            "B (有望)" if s >= 35 else
            "C (標準)" if s >= 15 else
            "D (低)"
        )

        result.append({
            "施設名": f["company_name"],
            "所在地": f["location"],
            "求人数": f["job_count"],
            "募集職種": " / ".join(list(dict.fromkeys(f["job_titles"]))[:5]),
            "スコア": f["best_score"],
            "優先度": f["priority"],
            "判定理由": "; ".join(sorted(f["all_reasons"])[:8]),
            "求人番号": ", ".join(f["job_numbers"][:3]),
            "取得日": f["retrieved_at"],
            "概要": (f["descriptions"][0][:150] + "...") if f["descriptions"] else "",
        })

    result.sort(key=lambda x: x["スコア"], reverse=True)
    return result


def normalize_company_name(name):
    """施設名を正規化して名寄せに使う。"""
    name = name.strip()
    name = re.sub(r"[\s　]+", " ", name)
    # 法人格を統一
    name = re.sub(r"（医）|[(（]医療法人[)）]", "医療法人", name)
    name = re.sub(r"（社）|[(（]社会福祉法人[)）]", "社会福祉法人", name)
    name = re.sub(r"（福）", "社会福祉法人", name)
    return name


# ── CSV / Excel 出力 ──────────────────────────────

def export_csv(facilities, filepath):
    """施設リストを CSV 出力する。"""
    if not facilities:
        print("  出力対象が0件です。")
        return

    fieldnames = ["優先度", "スコア", "施設名", "所在地", "求人数",
                  "募集職種", "判定理由", "求人番号", "取得日", "概要"]

    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for facility in facilities:
            writer.writerow(facility)

    print(f"\n  CSV 出力: {filepath}")
    print(f"  合計: {len(facilities)} 施設")


def export_excel(facilities, filepath):
    """施設リストを Excel 出力する（openpyxl が必要）。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("  [警告] openpyxl がないため Excel 出力をスキップ。CSV を使用してください。")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "見込み施設リスト"

    # ヘッダースタイル
    header_fill = PatternFill(start_color="0068B7", end_color="0068B7", fill_type="solid")
    header_font = Font(name="Yu Gothic", bold=True, color="FFFFFF", size=10)
    cell_font = Font(name="Yu Gothic", size=10)
    thin_border = Border(
        left=Side(style="thin", color="D4DDE8"),
        right=Side(style="thin", color="D4DDE8"),
        top=Side(style="thin", color="D4DDE8"),
        bottom=Side(style="thin", color="D4DDE8"),
    )

    # 優先度別の背景色
    priority_fills = {
        "A (最優先)": PatternFill(start_color="FFF2CC", fill_type="solid"),
        "B (有望)": PatternFill(start_color="E8F4FD", fill_type="solid"),
        "C (標準)": PatternFill(start_color="F5F5F5", fill_type="solid"),
        "D (低)": PatternFill(start_color="FFFFFF", fill_type="solid"),
    }

    headers = ["優先度", "スコア", "施設名", "所在地", "求人数",
               "募集職種", "判定理由", "求人番号", "取得日", "アクション"]
    col_widths = [12, 8, 30, 20, 8, 25, 35, 18, 12, 15]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col)].width = width

    for row_idx, facility in enumerate(facilities, 2):
        priority = facility.get("優先度", "D (低)")
        fill = priority_fills.get(priority, priority_fills["D (低)"])

        values = [
            facility.get("優先度", ""),
            facility.get("スコア", 0),
            facility.get("施設名", ""),
            facility.get("所在地", ""),
            facility.get("求人数", 0),
            facility.get("募集職種", ""),
            facility.get("判定理由", ""),
            facility.get("求人番号", ""),
            facility.get("取得日", ""),
            "",  # アクション列（手動記入用）
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = cell_font
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(col in [6, 7]))

    # フィルター
    ws.auto_filter.ref = f"A1:J{len(facilities) + 1}"

    # フリーズペイン
    ws.freeze_panes = "A2"

    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    print(f"  Excel 出力: {filepath}")


def print_summary(facilities):
    """結果のサマリーを表示する。"""
    total = len(facilities)
    by_priority = defaultdict(int)
    for f in facilities:
        by_priority[f["優先度"]] += 1

    print("\n" + "=" * 60)
    print("  YUMEHO 見込み施設リスト — サマリー")
    print("=" * 60)
    print(f"\n  合計: {total} 施設\n")
    for p in ["A (最優先)", "B (有望)", "C (標準)", "D (低)"]:
        count = by_priority.get(p, 0)
        bar = "█" * count + "░" * max(0, 20 - count)
        print(f"  {p:12s}  {bar} {count} 施設")

    # 上位10施設
    if facilities:
        print(f"\n{'─' * 60}")
        print("  上位10施設:")
        print(f"{'─' * 60}")
        for i, f in enumerate(facilities[:10], 1):
            print(f"  {i:2d}. [{f['優先度'][:1]}] {f['施設名'][:25]:25s} "
                  f"スコア:{f['スコア']:3d}  求人:{f['求人数']}件  {f['所在地'][:15]}")

    print()


# ── CSV 読み込み（再分析用） ──────────────────────

def load_csv(filepath):
    """既存の CSV を読み込んで再スコアリング用のリストに変換する。"""
    listings = []
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            listing = {
                "job_title": row.get("募集職種", row.get("job_title", "")),
                "company_name": row.get("施設名", row.get("company_name", "")),
                "location": row.get("所在地", row.get("location", "")),
                "description": row.get("概要", row.get("description", "")),
                "requirements": row.get("requirements", ""),
                "job_number": row.get("求人番号", row.get("job_number", "")),
                "retrieved_at": row.get("取得日", row.get("retrieved_at", "")),
            }
            listings.append(listing)
    return listings


# ── メイン処理 ──────────────────────────────────

def cmd_search(args):
    """ハローワークを検索して施設リストを作成する。"""
    areas = []
    keywords = []

    # 引数パース
    i = 0
    while i < len(args):
        if args[i] in ("--area", "-a") and i + 1 < len(args):
            areas.append(args[i + 1])
            i += 2
        elif args[i] in ("--keyword", "-k") and i + 1 < len(args):
            keywords.append(args[i + 1])
            i += 2
        elif args[i] in ("--all-areas",):
            areas = list(AREA_CODES.keys())
            i += 1
        else:
            i += 1

    if not areas:
        areas = DEFAULT_AREAS
    if not keywords:
        keywords = [
            "理学療法士 回復期",
            "理学療法士 リハビリ",
            "作業療法士 リハビリ",
            "介護職員 リハビリ",
            "機能訓練指導員",
            "歩行訓練 介護",
        ]

    print("=" * 60)
    print("  YUMEHO 見込み施設抽出ツール")
    print("  ハローワークインターネットサービス検索")
    print("=" * 60)
    print(f"\n  対象エリア: {', '.join(areas)}")
    print(f"  検索キーワード: {len(keywords)} パターン")
    print(f"  リクエスト間隔: {REQUEST_INTERVAL} 秒")

    searcher = HelloWorkSearcher()
    total = 0

    for area in areas:
        print(f"\n{'─' * 40}")
        print(f"  エリア: {area}")
        print(f"{'─' * 40}")

        for kw in keywords:
            count = searcher.search(kw, area_name=area, max_pages=3)
            total += count
            time.sleep(REQUEST_INTERVAL)

    print(f"\n  取得完了: 全 {total} 件の求人情報")

    # 名寄せ・スコアリング
    print("\n  施設の名寄せ・スコアリング中...")
    facilities = deduplicate_and_aggregate(searcher.results)

    # 出力
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = os.path.join(OUTPUT_DIR, f"yumeho_leads_{timestamp}.csv")
    xlsx_path = os.path.join(OUTPUT_DIR, f"yumeho_leads_{timestamp}.xlsx")

    export_csv(facilities, csv_path)
    export_excel(facilities, xlsx_path)
    print_summary(facilities)

    return facilities


def cmd_analyze(args):
    """既存の CSV を読み込んで再スコアリングする。"""
    if not args:
        print("  使い方: python3 hellowork-lead-finder.py analyze <CSVファイルパス>")
        return

    filepath = args[0]
    if not os.path.exists(filepath):
        print(f"  [エラー] ファイルが見つかりません: {filepath}")
        return

    print("=" * 60)
    print("  YUMEHO 見込み施設 — 再スコアリング")
    print("=" * 60)

    listings = load_csv(filepath)
    print(f"\n  読み込み: {len(listings)} 件")

    facilities = deduplicate_and_aggregate(listings)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = os.path.join(OUTPUT_DIR, f"yumeho_leads_rescored_{timestamp}.csv")
    xlsx_path = os.path.join(OUTPUT_DIR, f"yumeho_leads_rescored_{timestamp}.xlsx")

    export_csv(facilities, csv_path)
    export_excel(facilities, xlsx_path)
    print_summary(facilities)


def cmd_demo(args):
    """デモ用: サンプルデータでスコアリングの動作を確認する。"""
    print("=" * 60)
    print("  YUMEHO 見込み施設抽出ツール — デモモード")
    print("=" * 60)

    sample_listings = [
        {
            "job_title": "理学療法士（回復期リハビリテーション病棟）",
            "company_name": "医療法人社団 リハビリテーション花の丘病院",
            "location": "東京都世田谷区",
            "description": "回復期リハビリテーション病棟での歩行訓練を中心としたリハビリ業務。"
                           "急募。転倒防止に配慮した安全な環境で、患者様の歩行機能回復をサポート。"
                           "身体介助を含む。増員のため2名募集。",
            "requirements": "理学療法士免許必須",
            "job_number": "13070-12345678",
            "retrieved_at": "2026-04-12",
        },
        {
            "job_title": "介護職員（デイサービス）",
            "company_name": "社会福祉法人 さくら福祉会",
            "location": "神奈川県横浜市",
            "description": "通所介護施設での介護業務全般。歩行介助、移乗介助、入浴介助など。"
                           "リハビリプログラムの補助。急募。3名募集。未経験可。",
            "requirements": "介護職員初任者研修以上",
            "job_number": "14010-23456789",
            "retrieved_at": "2026-04-12",
        },
        {
            "job_title": "作業療法士",
            "company_name": "介護老人保健施設 グリーンヒルズ",
            "location": "埼玉県さいたま市",
            "description": "老健での機能訓練業務。入所者様の日常生活動作の改善、"
                           "歩行訓練、認知機能訓練など。欠員補充のため募集。",
            "requirements": "作業療法士免許必須",
            "job_number": "11010-34567890",
            "retrieved_at": "2026-04-12",
        },
        {
            "job_title": "機能訓練指導員",
            "company_name": "有限会社 ケアプラス",
            "location": "千葉県船橋市",
            "description": "デイサービスでの機能訓練業務。個別機能訓練計画の作成と実施。",
            "requirements": "柔道整復師、あん摩マッサージ指圧師等",
            "job_number": "12040-45678901",
            "retrieved_at": "2026-04-12",
        },
        {
            "job_title": "理学療法士（リハビリテーション科）",
            "company_name": "医療法人社団 リハビリテーション花の丘病院",
            "location": "東京都世田谷区",
            "description": "外来リハビリテーション。脳血管疾患、整形外科疾患の患者様の"
                           "歩行訓練・ADL訓練。転倒リスクの高い患者様への対応。",
            "requirements": "理学療法士免許必須、3年以上経験者優遇",
            "job_number": "13070-12345679",
            "retrieved_at": "2026-04-12",
        },
        {
            "job_title": "介護職員",
            "company_name": "特別養護老人ホーム やすらぎの里",
            "location": "大阪府大阪市",
            "description": "特養での介護業務。食事・入浴・排泄介助。移乗介助。"
                           "夜勤あり。人手不足のため随時募集。ブランク可。",
            "requirements": "資格不問、未経験可",
            "job_number": "27010-56789012",
            "retrieved_at": "2026-04-12",
        },
    ]

    print(f"\n  サンプルデータ: {len(sample_listings)} 件の求人")

    facilities = deduplicate_and_aggregate(sample_listings)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = os.path.join(OUTPUT_DIR, f"yumeho_leads_demo_{timestamp}.csv")
    xlsx_path = os.path.join(OUTPUT_DIR, f"yumeho_leads_demo_{timestamp}.xlsx")

    export_csv(facilities, csv_path)
    export_excel(facilities, xlsx_path)
    print_summary(facilities)


def main():
    if len(sys.argv) < 2:
        print("""
YUMEHO 見込み施設抽出ツール
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

使い方:
  python3 hellowork-lead-finder.py search                      # デフォルト検索（首都圏+関西+東海）
  python3 hellowork-lead-finder.py search --area 東京都         # エリア指定
  python3 hellowork-lead-finder.py search --area 大阪府 --area 兵庫県  # 複数エリア
  python3 hellowork-lead-finder.py search --all-areas          # 全国検索
  python3 hellowork-lead-finder.py search --keyword "歩行訓練"  # キーワード指定
  python3 hellowork-lead-finder.py analyze leads.csv           # 既存CSV再スコアリング
  python3 hellowork-lead-finder.py demo                        # デモ（サンプルデータ）

出力:
  leads/yumeho_leads_YYYYMMDD_HHMMSS.csv   — CSV（Excel で開ける）
  leads/yumeho_leads_YYYYMMDD_HHMMSS.xlsx  — Excel（色付き・フィルター付き）

スコアリング基準:
  - 施設種別（回復期リハビリ病院、老健、デイケア等）: +20点
  - 職種（PT/OT: +15, 介護職: +10, リハビリ関連: +12）
  - 歩行リハビリ直接言及: +25点
  - 転倒関連言及: +15点
  - 急募: +15点
  - 複数名募集: +10-50点
  - その他シグナル（人手不足、欠員補充等）: 各+5点
  - 複数職種で求人: +10-20点

優先度:
  A (最優先): 60点以上 — 即アプローチ
  B (有望):   35点以上 — 優先的にアプローチ
  C (標準):   15点以上 — リスト保持
  D (低):     15点未満 — 見送り
        """)
        return

    command = sys.argv[1]
    args = sys.argv[2:]

    if command == "search":
        cmd_search(args)
    elif command == "analyze":
        cmd_analyze(args)
    elif command == "demo":
        cmd_demo(args)
    else:
        print(f"  [エラー] 不明なコマンド: {command}")
        print("  search / analyze / demo のいずれかを指定してください。")


if __name__ == "__main__":
    main()
