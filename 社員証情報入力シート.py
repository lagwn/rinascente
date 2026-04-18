import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "社員証情報入力"

NUM_PERSONS = 6

# --- スタイル定義 ---
title_font = Font(name="Yu Gothic", size=16, bold=True)
header_font = Font(name="Yu Gothic", size=11, bold=True, color="FFFFFF")
input_font = Font(name="Yu Gothic", size=11)
example_font = Font(name="Yu Gothic", size=9, color="888888")
note_font = Font(name="Yu Gothic", size=10, color="FF0000")
label_font = Font(name="Yu Gothic", size=11, bold=True)
person_header_font = Font(name="Yu Gothic", size=12, bold=True, color="FFFFFF")

header_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
label_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
input_fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
person_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

# --- 列幅設定 ---
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 22
for i in range(NUM_PERSONS):
    col_letter = get_column_letter(3 + i)
    ws.column_dimensions[col_letter].width = 30
last_col_letter = get_column_letter(3 + NUM_PERSONS)
ws.column_dimensions[last_col_letter].width = 2

end_col = get_column_letter(2 + NUM_PERSONS)

# --- タイトル ---
ws.merge_cells(f"B2:{end_col}2")
title_cell = ws["B2"]
title_cell.value = "社員証作成 情報入力シート"
title_cell.font = title_font
title_cell.alignment = center_align

ws.merge_cells(f"B3:{end_col}3")
subtitle_cell = ws["B3"]
subtitle_cell.value = "※ 黄色のセルに情報をご入力ください ／ 顔写真は別途ご提出ください"
subtitle_cell.font = note_font
subtitle_cell.alignment = center_align

# --- ヘッダー行 ---
row = 5

item_header = ws.cell(row=row, column=2, value="項目")
item_header.font = header_font
item_header.fill = header_fill
item_header.alignment = center_align
item_header.border = thin_border

for i in range(NUM_PERSONS):
    col = 3 + i
    cell = ws.cell(row=row, column=col, value=f"{i+1}人目")
    cell.font = person_header_font
    cell.fill = person_header_fill
    cell.alignment = center_align
    cell.border = thin_border

# --- 入力項目 ---
items = [
    "肩書き（役職）",
    "氏名",
    "ふりがな",
    "顔写真ファイル名",
]

for i, label in enumerate(items):
    r = row + 1 + i
    ws.row_dimensions[r].height = 38

    # 項目名
    label_cell = ws.cell(row=r, column=2, value=label)
    label_cell.font = label_font
    label_cell.alignment = center_align
    label_cell.border = thin_border
    label_cell.fill = label_fill

    # 各人の入力欄
    for p in range(NUM_PERSONS):
        col = 3 + p
        if "顔写真" in label and p == 0:
            input_cell = ws.cell(row=r, column=col, value="例）yamada_taro.jpg")
            input_cell.font = example_font
        else:
            input_cell = ws.cell(row=r, column=col, value="")
            input_cell.font = input_font
        input_cell.alignment = left_align
        input_cell.border = thin_border
        input_cell.fill = input_fill
        input_cell.number_format = "@"

# --- 注意書き ---
note_start = row + 1 + len(items) + 1
ws.merge_cells(f"B{note_start}:{end_col}{note_start}")
note_cell = ws.cell(row=note_start, column=2)
note_cell.value = "【ご注意】"
note_cell.font = Font(name="Yu Gothic", size=11, bold=True, color="FF0000")

notes = [
    "・ 入力が完了しましたら、本ファイルと顔写真データをあわせてご返送ください。",
    "・ 氏名等は社員証に印刷されるとおりに正確にご入力ください。",
    "・ 顔写真ファイル名欄には、提出いただく写真のファイル名をご記入ください。",
    "・ 不明点がございましたらお気軽にお問い合わせください。",
]
for j, note in enumerate(notes):
    r = note_start + 1 + j
    ws.merge_cells(f"B{r}:{end_col}{r}")
    cell = ws.cell(row=r, column=2, value=note)
    cell.font = Font(name="Yu Gothic", size=10)
    cell.alignment = left_align

# --- 顔写真についての案内 ---
photo_start = note_start + 1 + len(notes) + 1
ws.merge_cells(f"B{photo_start}:{end_col}{photo_start}")
photo_title = ws.cell(row=photo_start, column=2)
photo_title.value = "【顔写真データについて】"
photo_title.font = Font(name="Yu Gothic", size=11, bold=True, color="2B5797")

photo_notes = [
    "■ カードサイズ（86mm × 54mm）に印刷するため、下記の条件を満たす写真をご用意ください。",
    "",
    "  ● 解像度：300dpi 以上（印刷品質を確保するため）",
    "  ● 推奨画像サイズ：600 × 600 ピクセル以上（目安）",
    "      ※ 顔写真の印刷領域を約25mm × 30mmとした場合、300dpiで約300×354ピクセル以上必要です。",
    "      ※ トリミング調整の余裕を考慮し、600×600ピクセル以上を推奨します。",
    "  ● ファイル形式：JPEG（.jpg）推奨  ※ PNG(.png) も可",
    "  ● ファイルサイズ：1MB 以上推奨（圧縮しすぎると画質が劣化します）",
    "  ● 撮影条件：正面向き、無帽、背景は白または無地、胸から上が写っているもの",
    "",
    "  ※ スマートフォンで撮影する場合は、明るい場所で撮影し、最高画質設定でお願いします。",
    "  ※ 解像度が不足している場合、印刷時にぼやけた仕上がりになります。",
]
for j, line in enumerate(photo_notes):
    r = photo_start + 1 + j
    ws.merge_cells(f"B{r}:{end_col}{r}")
    cell = ws.cell(row=r, column=2, value=line)
    cell.font = Font(name="Yu Gothic", size=10)
    cell.alignment = left_align

# --- 印刷設定 ---
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.page_setup.orientation = "landscape"

# --- 保存 ---
output_path = "/Users/naoya/Desktop/クライアント1119/真誠会/プロトタイプ/社員証情報入力シート.xlsx"
wb.save(output_path)
print(f"保存完了: {output_path}")
