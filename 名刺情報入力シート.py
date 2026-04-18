import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "名刺情報入力"

NUM_PERSONS = 6

# --- スタイル定義 ---
title_font = Font(name="Yu Gothic", size=16, bold=True)
header_font = Font(name="Yu Gothic", size=11, bold=True, color="FFFFFF")
input_font = Font(name="Yu Gothic", size=11)
example_font = Font(name="Yu Gothic", size=9, color="888888")
note_font = Font(name="Yu Gothic", size=10, color="FF0000")
label_font = Font(name="Yu Gothic", size=11, bold=True)
person_header_font = Font(name="Yu Gothic", size=12, bold=True, color="FFFFFF")

header_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
label_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
input_fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
person_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

# --- 列幅設定 ---
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 22  # 項目列
for i in range(NUM_PERSONS):
    col_letter = get_column_letter(3 + i)  # C, D, E, F, G, H
    ws.column_dimensions[col_letter].width = 30
last_col_letter = get_column_letter(3 + NUM_PERSONS)
ws.column_dimensions[last_col_letter].width = 2

# --- タイトル ---
end_col = get_column_letter(2 + NUM_PERSONS)
ws.merge_cells(f"B2:{end_col}2")
title_cell = ws["B2"]
title_cell.value = "名刺作成 情報入力シート"
title_cell.font = title_font
title_cell.alignment = center_align

ws.merge_cells(f"B3:{end_col}3")
subtitle_cell = ws["B3"]
subtitle_cell.value = "※ 黄色のセルに情報をご入力ください"
subtitle_cell.font = note_font
subtitle_cell.alignment = center_align

# --- ヘッダー行（人数分） ---
row = 5

# 項目ヘッダー
item_header = ws.cell(row=row, column=2, value="項目")
item_header.font = header_font
item_header.fill = header_fill
item_header.alignment = center_align
item_header.border = thin_border

for i in range(NUM_PERSONS):
    col = 3 + i
    cell = ws.cell(row=row, column=col, value=f"{i+1}人目")
    cell.font = person_header_font
    cell.fill = person_header_fill
    cell.alignment = center_align
    cell.border = thin_border

# --- 入力項目定義 ---
items = [
    "肩書き",
    "氏名",
    "ふりがな",
    "郵便番号",
    "住所（1行目）",
    "住所（2行目）",
    "電話番号",
    "FAX番号",
    "個人携帯番号",
    "メールアドレス\n（ご希望アドレス）",
]

for i, label in enumerate(items):
    r = row + 1 + i
    ws.row_dimensions[r].height = 38 if "メール" not in label else 48

    # 項目名
    label_cell = ws.cell(row=r, column=2, value=label)
    label_cell.font = label_font
    label_cell.alignment = center_align
    label_cell.border = thin_border
    label_cell.fill = label_fill

    # 各人の入力欄
    for p in range(NUM_PERSONS):
        col = 3 + p
        # メールアドレス行の1人目だけ入力例を薄く表示
        if "メール" in label and p == 0:
            input_cell = ws.cell(row=r, column=col, value="例）kawahira@rinascentes.co.jp")
            input_cell.font = example_font
        else:
            input_cell = ws.cell(row=r, column=col, value="")
            input_cell.font = input_font
        input_cell.alignment = left_align
        input_cell.border = thin_border
        input_cell.fill = input_fill
        input_cell.number_format = "@"  # テキスト書式（先頭0保持）

# --- 注意書き ---
note_start = row + 1 + len(items) + 1
ws.merge_cells(f"B{note_start}:{end_col}{note_start}")
note_cell = ws.cell(row=note_start, column=2)
note_cell.value = "【ご注意】"
note_cell.font = Font(name="Yu Gothic", size=11, bold=True, color="FF0000")

notes = [
    "・ 入力が完了しましたら、本ファイルをご返送ください。",
    "・ メールアドレスは新規作成となりますので、ご希望のアドレスをご記入ください。",
    "・ 氏名・住所等は名刺に印刷されるとおりに正確にご入力ください。",
    "・ 住所・電話番号等が全員共通の場合は、1人目のみご記入いただければ結構です。",
    "・ 不明点がございましたらお気軽にお問い合わせください。",
]
for j, note in enumerate(notes):
    r = note_start + 1 + j
    ws.merge_cells(f"B{r}:{end_col}{r}")
    cell = ws.cell(row=r, column=2, value=note)
    cell.font = Font(name="Yu Gothic", size=10)
    cell.alignment = left_align

# --- 印刷設定 ---
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.page_setup.orientation = "landscape"

# --- 保存 ---
output_path = "/Users/naoya/Desktop/クライアント1119/真誠会/プロトタイプ/名刺情報入力シート.xlsx"
wb.save(output_path)
print(f"保存完了: {output_path}")
